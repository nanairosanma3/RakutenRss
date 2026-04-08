import time
import os
import xlwings as xw
from pathlib import Path
from datetime import datetime, timezone, timedelta
import subprocess
import hashlib
import json
import re
import openpyxl as pyxl
from multiprocessing import Process
from datetime import time as datetime_time
from dotenv import load_dotenv


load_dotenv()


#####参考サイト：https://zenn.dev/ririkabu/articles/b7adbbd3012eea


class OrderBookMonitor:
    def __init__(self, excel_path, code_list, json_base_path):
        self.excel_path     = excel_path
        self.code_list      = code_list
        self.json_base_path = Path(json_base_path)
        self.pid            = None # Exel を起動するprocess id
        self.previous_hashes = {}
        self.headers        = []
#        self.data_queue     = MPQueue # Multiprocessing Queue
        self.stop_flag      = False
        self.reader_process = None # データ取得プロセスの参照

        # `data_range` を code_list の範囲から動的に設定
        start_row = 2
        end_row   = start_row + len( code_list ) - 1
        self.data_range = f" B{ start_row } : ES{ end_row } "




    # excelファイルを新規作成し、アクティブなシートのA1から
    # 市況情報関数のヘッダー行を関数ごとに表示
    def create_excel(self):
        wb = pyxl.Workbook()
        ws = wb.active
        ws["A1"].value = "=RssMarketHeader(4)" # 1 = 国内株式、2 = 先物、3 = 指数、4 = 為替

        #市場データを取得するセルの設定
        # row = 行, col = column = 列
        # for i in range( A, B ) --> i はAから( B - 1 )まで変化する
        for row in range(2, 30):
            for col in range(2, 25):
                cell = ws.cell( row = row, column = col )
                # RssMaket("銘柄コード", "取得項目"), pyxl.utils.get_column_letter = 数字をアルファベットに
                cell.value = f"=RssFXMarket($A{row}, {pyxl.utils.get_column_letter(col)}$1 )"

        # 銘柄コードの設定
        for i, code in enumerate(self.code_list):
            ws[ f"A{ i + 2 }" ].value = str(code) 

        wb.save(self.excel_path)



    def get_path_to_xl(self) -> Path:
        try:
            subprocess_rtn = subprocess.run("assoc .xlsx", shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            assoc_to = re.search(r"Excel\.Sheet\.\d+", subprocess_rtn.stdout.decode("utf-8")).group()


            subprocess_rtn = subprocess.run(f"ftype {assoc_to}", shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            xl_path = re.search(r"C:.*EXCEL\.EXE", subprocess_rtn.stdout.decode("utf-8")).group()

###            print("Excelのパス:", xl_path) # デバック用出力
            return Path(xl_path)

        except AttributeError as e:
            raise FileNotFoundError("Excelのパスが取得できませんでした。Excelがインストールされているか確認してください。") from e

    
    # /x : 新しいExcelを別のウィンドウで起動する
    # Popen : run と違い並列実行される。Popen で起動したものが終了
    #       　しているかどうかに関わらず、Popen の後に記述されたコードは実行される。
    def add_xl_app(self) -> xw.App:
        try:
            excel_path = self.get_path_to_xl()
            command = f'"{str(excel_path)}" /x "{self.excel_path}"'
###            print("実行コマンド:", command) # デバック用出力
            proc = subprocess.Popen(command)
            time.sleep(1) # 初期待機

            for _ in range(10):
                try:
                    xl_app = xw.apps[proc.pid]
                    print("PID", proc.pid, xw.apps.keys(), "アプリケーションが正常に起動しました。")
                    return xl_app, proc.pid
                except KeyError:
###                    print("PID確認中...") # デバック用出力
                    time.sleep(0.5)

            proc.terminate()
            raise RuntimeError("Excelアプリケーションが正常に起動しませんでした。")
        except Exception as e:
            print("エラー内容:", e)
            raise RuntimeError("Excelの起動に失敗しました。") from e


    def initialize_excel(self):
        app, self.pid = self.add_xl_app()
        time.sleep(2)
        wb = app.books.active
        sht = wb.sheets[0]

        # ヘッダーの取得
        self.headers = sht.range("B1:T1").value
###        print(f"監視する銘柄コード: {self.code_list}")
        print("Excelシートの初期化が完了しました。")

    def get_current_timestamp(self):
        jst = timezone(timedelta(hours=9))
        return datetime.now(jst).strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]

    def calculate_hash(self,data):
        data_without_timestamp = {k:v for k, v in data.items() if k != "timestamp"}
        data_string = json.dumps(data_without_timestamp, sort_keys=True).encode("utf-8")
        return hashlib.md5(data_string).hexdigest()

    def watch_loop(self):
        
        ws = xw.sheets[0]
        time.sleep(5)
        print("監視員、着任いたしました！監視を開始します！！")
        while True:

            current_data = {"stock_code": ws.range("A2").value, "price": ws.range("F2").value}
            current_hash = self.calculate_hash(current_data)

            if self.previous_hashes != current_hash:

                print("現在値:", current_data["price"])

                self.previous_hashes = current_hash

                time.sleep(1)


    def monitor(self):
        try:
            self.create_excel()
            self.initialize_excel()
            self.watch_loop()
        except KeyboardInterrupt:
            print("停止処理を開始しています...")
            self.stop_flag = True # 停止フラグを設定してループを終了させる
        finally:

            print("監視を停止しました。")



if __name__ == "__main__":
    EXCEL_PATH = os.getenv("TEST_EXCEL_PATH") 
    INIT_CODE = ["USD/JPY", "EUR/JPY", "GBP/JPY", "AUD/JPY", "NZD/JPY", "ZAR/JPY", "CAD/JPY", "CHF/JPY", "N225","N225.FUT01.OS" ]
    #INIT_CODE = get_watchlist_codes()
    JSON_BASE_PATH = os.getenv("JSON_BASE_PATH")

    target_time = datetime_time(6, 0)

    while True:
        current_time = datetime.now().time()
        if current_time >= target_time and current_time:
            print(f"現在時刻: {current_time}. 処理を開始します。 ")
            # モニターのインスタンス作成
            monitor = OrderBookMonitor(EXCEL_PATH, INIT_CODE, JSON_BASE_PATH)
            monitor.monitor()
            break # メイン処理が終了したらループを抜ける
        else:
            print(f"現在時刻: {current_time}. 開始時刻 {target_time} まで待機中...")
            time.sleep(60)





#####参考ここまで
